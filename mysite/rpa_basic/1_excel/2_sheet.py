from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet()  # 새로운 sheet 기분 이름으로 생성
ws.title = "MySheet"
ws.sheet_properties.tabColor = "ff66ff"

ws1 = wb.create_sheet("YourSheet")
ws2 = wb.create_sheet("NewSheet", 2)

new_ws = wb["NewSheet"]

print(wb.sheetnames)

new_ws["A1"] = "TEst"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")
